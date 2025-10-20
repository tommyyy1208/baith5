import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
import time
import re

# ==========================
# ĐỌC FILE CLASS ĐA NGÔN NGỮ - TỐI ƯU
# ==========================
def load_classes_from_file(filename="classes.txt"):
    if not os.path.exists(filename):
        messagebox.showerror("Error", f"File {filename} not found.")
        return {}, [], {}

    with open(filename, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    lang_dict = {}
    class_mapping = {}
    current_lang = None
    current_class_index = 0
    parent_classes = {}
    child_classes = []

    for line in lines:
        if re.match(r"^\d+\.\s", line):
            lang_name = line.split(". ", 1)[1]
            current_lang = lang_name
            lang_dict[current_lang] = []
            current_class_index = 0
        elif re.match(r"^\d+\.\d+\.\s", line) and current_lang:
            cls_name = line.split(". ", 1)[1]
            lang_dict[current_lang].append(cls_name)
            
            if current_class_index not in class_mapping:
                class_mapping[current_class_index] = {}
            class_mapping[current_class_index][current_lang] = cls_name
            parent_classes[current_class_index] = cls_name
            current_class_index += 1
        elif re.match(r"^\d+\.\d+\.\d+\.\s", line) and current_lang:
            cls_name = line.split(". ", 1)[1]
            lang_dict[current_lang].append(cls_name)
            
            if current_class_index not in class_mapping:
                class_mapping[current_class_index] = {}
            class_mapping[current_class_index][current_lang] = cls_name
            child_classes.append(current_class_index)
            current_class_index += 1

    languages = list(lang_dict.keys())
    return lang_dict, languages, class_mapping, parent_classes, child_classes

# Hàm validate chỉ cho phép nhập số
def validate_number_input(new_value):
    return new_value == "" or new_value.isdigit()

# ==========================
# ỨNG DỤNG CHÍNH - ĐÃ TỐI ƯU
# ==========================
class CounterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Annotation Counter Tool")
        self.root.geometry("1400x900")

        # Pre-load data
        self.class_sets, self.languages, self.class_mapping, self.parent_classes, self.child_classes = load_classes_from_file()
        if not self.languages:
            messagebox.showerror("Error", "No class definitions found in classes.txt")
            self.root.destroy()
            return

        # ĐẢM BẢO class_indexes ĐƯỢC KHỞI TẠO
        self.class_indexes = list(self.class_mapping.keys())
        
        self.current_language = tk.StringVar(value=self.languages[0])
        
        # Cache để tăng tốc độ
        self._language_cache = {}
        self._column_mapping_cache = None
        
        # Khởi tạo counts_all
        self.counts_all = {}
        for index in self.class_indexes:
            self.counts_all[index] = tk.IntVar(value=0)

        self.count_labels = {}
        self.entry_widgets = {}
        self.total_work_time = 0.0
        self.session_start = None
        self.is_paused = True

        # Đăng ký hàm validate
        self.vcmd = (self.root.register(validate_number_input), '%P')

        self.setup_ui()
        self.update_timer_display()

    # ----------------------------
    # GIAO DIỆN CHÍNH - TỐI ƯU
    # ----------------------------
    def setup_ui(self):
        # Sử dụng frame đơn giản hơn, không dùng canvas scroll cho phần chính
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill="both", expand=True)

        # Phần trên cùng - được giữ nguyên
        self.setup_top_controls(main_frame)
        
        # Phần class - tối ưu hiển thị
        self.setup_class_display(main_frame)

    def setup_top_controls(self, parent):
        """Thiết lập các control phía trên - ĐÃ ĐIỀU CHỈNH LAYOUT"""
        # Tạo frame chính cho các control trên cùng
        top_main_frame = ttk.Frame(parent)
        top_main_frame.pack(pady=10, fill="x")
        
        # Hàng 1: Dataset Name và Display Language
        row1_frame = ttk.Frame(top_main_frame)
        row1_frame.pack(fill="x", pady=5)
        
        # Dataset Name
        ttk.Label(row1_frame, text="Dataset Name:").pack(side="left")
        self.dataset_name_entry = ttk.Entry(row1_frame, width=30)
        self.dataset_name_entry.pack(side="left", padx=5)

        # Display Language
        ttk.Label(row1_frame, text="Display Language:").pack(side="left", padx=(20, 5))
        self.lang_menu = ttk.OptionMenu(
            row1_frame, self.current_language, self.languages[0], *self.languages, 
            command=self.update_language
        )
        self.lang_menu.pack(side="left", padx=(0, 20))

        # Working Time - ĐƯA LÊN CÙNG HÀNG
        self.timer_label = ttk.Label(row1_frame, text="Working Time: 00:00:00", 
                                   font=("Arial", 11, "bold"))
        self.timer_label.pack(side="left", padx=(20, 0))

        # Hàng 2: Nút Play/Pause và Save/Load
        row2_frame = ttk.Frame(top_main_frame)
        row2_frame.pack(fill="x", pady=5)
        
        # Nút Play / Pause
        ttk.Label(row2_frame, text="Timer Controls:").pack(side="left")
        ttk.Button(row2_frame, text="▶ Play", 
                  command=self.start_timer).pack(side="left", padx=5)
        ttk.Button(row2_frame, text="⏸ Pause", 
                  command=self.pause_timer).pack(side="left", padx=5)

        # Nút Save và Load - ĐƯA LÊN CÙNG HÀNG
        ttk.Label(row2_frame, text="Data Management:").pack(side="left", padx=(40, 5))
        ttk.Button(row2_frame, text="💾 Save to Excel", 
                  command=self.save_to_excel).pack(side="left", padx=5)
        ttk.Button(row2_frame, text="📂 Load from Excel", 
                  command=self.load_from_excel).pack(side="left", padx=5)

    def setup_class_display(self, parent):
        """Thiết lập hiển thị class với canvas scroll - tối ưu"""
        # Tạo container với scrollbar
        container = ttk.Frame(parent)
        container.pack(fill="both", expand=True, pady=10)
        
        # Tạo canvas và scrollbar
        self.canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        # Configure canvas scrolling
        self.scrollable_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        # Pack với tỷ lệ phù hợp
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Bind sự kiện scroll
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.scrollable_frame.bind("<MouseWheel>", self._on_mousewheel)

        # Frame chính cho class
        self.class_frame = ttk.Frame(self.scrollable_frame)
        self.class_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Pre-render layout
        self._precompute_column_mapping()

    def _precompute_column_mapping(self):
        """Tính toán trước column mapping để tăng tốc"""
        if self._column_mapping_cache is None:
            self._column_mapping_cache = {}
            for parent_index, parent_name in self.parent_classes.items():
                # Kiểm tra cache cho ngôn ngữ đầu tiên
                current_parent_name = self.class_mapping[parent_index].get(self.languages[0], parent_name)
                
                if any(x in current_parent_name for x in ["車線", "Lane", "Làn đường"]):
                    self._column_mapping_cache[parent_index] = "LEFT"
                elif any(x in current_parent_name for x in ["横断", "Crosswalk", "Vạch sang đường"]):
                    self._column_mapping_cache[parent_index] = "LEFT"
                elif any(x in current_parent_name for x in ["センター", "Center", "Đường tâm"]):
                    self._column_mapping_cache[parent_index] = "CENTER"
                elif any(x in current_parent_name for x in ["交差点", "Intersection", "Giao lộ"]):
                    self._column_mapping_cache[parent_index] = "CENTER"
                elif any(x in current_parent_name for x in ["道路端", "Roadside", "Lề đường"]):
                    self._column_mapping_cache[parent_index] = "CENTER"
                elif any(x in current_parent_name for x in ["その他", "Others", "Khác"]):
                    self._column_mapping_cache[parent_index] = "RIGHT"
                else:
                    self._column_mapping_cache[parent_index] = "CENTER"

    def _on_frame_configure(self, event=None):
        """Cấu hình lại canvas khi frame thay đổi kích thước"""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_mousewheel(self, event):
        """Xử lý sự kiện scroll chuột"""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ----------------------------
    # CẬP NHẬT NGÔN NGỮ - TỐI ƯU
    # ----------------------------
    def update_language(self, lang):
        """Cập nhật ngôn ngữ với caching để tăng tốc"""
        # Kiểm tra cache
        if lang in self._language_cache:
            cached_data = self._language_cache[lang]
            self._rebuild_ui_from_cache(cached_data)
            return

        # Xóa widget cũ một cách hiệu quả
        for widget in self.class_frame.winfo_children():
            widget.destroy()

        self.count_labels.clear()
        self.entry_widgets.clear()

        # Tạo layout 3 cột
        columns_frame = ttk.Frame(self.class_frame)
        columns_frame.pack(fill="both", expand=True)
        
        left_column = ttk.Frame(columns_frame)
        left_column.pack(side="left", fill="both", expand=True, padx=5)
        
        center_column = ttk.Frame(columns_frame)
        center_column.pack(side="left", fill="both", expand=True, padx=5)
        
        right_column = ttk.Frame(columns_frame)
        right_column.pack(side="left", fill="both", expand=True, padx=5)

        # Cache để lưu kết quả
        cache_data = {
            'left': [],
            'center': [], 
            'right': []
        }

        # Phân bổ các cụm class
        for parent_index in sorted(self.parent_classes.keys()):
            if parent_index not in self._column_mapping_cache:
                continue
                
            parent_name = self.class_mapping[parent_index].get(lang, f"Parent_{parent_index}")
            column_type = self._column_mapping_cache[parent_index]
            
            if column_type == "LEFT":
                target_column = left_column
                cache_key = 'left'
            elif column_type == "CENTER":
                target_column = center_column
                cache_key = 'center'
            else:  # RIGHT
                target_column = right_column
                cache_key = 'right'

            # Tạo section
            section_frame = ttk.LabelFrame(target_column, text=parent_name, padding=5)
            section_frame.pack(fill="both", expand=True, pady=5)
            
            # Tìm class con
            child_indices = self._find_child_indices(parent_index)
            
            # Tạo các class con
            section_data = []
            for child_index in child_indices:
                if child_index in self.class_mapping:
                    class_name = self.class_mapping[child_index].get(lang, f"Class_{child_index}")
                    row_data = self._create_class_row(section_frame, class_name, child_index, column_type == "RIGHT")
                    section_data.append(row_data)
            
            cache_data[cache_key].append({
                'parent_name': parent_name,
                'children': section_data
            })

        # Lưu vào cache
        self._language_cache[lang] = cache_data

    def _find_child_indices(self, parent_index):
        """Tìm class con một cách hiệu quả"""
        child_indices = []
        sorted_parents = sorted(self.parent_classes.keys())
        
        # Tìm parent tiếp theo
        next_parent_index = None
        for next_idx in sorted_parents:
            if next_idx > parent_index:
                next_parent_index = next_idx
                break
        
        # Tìm class con trong khoảng
        if next_parent_index is not None:
            for idx in range(parent_index + 1, next_parent_index):
                if idx in self.class_mapping and idx not in self.parent_classes:
                    child_indices.append(idx)
        else:
            for idx in range(parent_index + 1, len(self.class_mapping)):
                if idx in self.class_mapping and idx not in self.parent_classes:
                    child_indices.append(idx)
        
        return child_indices

    def _create_class_row(self, frame, class_name, class_index, is_others_section=False):
        """Tạo một hàng class - tối ưu"""
        count_var = self.counts_all[class_index]

        row = ttk.Frame(frame)
        row.pack(fill="x", pady=1)

        # Hiển thị tên class - rộng hơn cho section "その他"
        label_width = 28 if is_others_section else 22
        ttk.Label(row, text=class_name, width=label_width, anchor="w").pack(side="left")
        
        # Frame điều khiển
        control_frame = ttk.Frame(row)
        control_frame.pack(side="right")
        
        # Nút và entry
        ttk.Button(control_frame, text="-", width=2,
                  command=lambda idx=class_index: self.decrement(idx)).pack(side="left")
        
        entry = ttk.Entry(control_frame, width=4, justify="center",
                         validate="key", validatecommand=self.vcmd)
        entry.pack(side="left", padx=2)
        entry.bind('<Return>', lambda e, idx=class_index: self.update_from_entry(idx))
        entry.bind('<FocusOut>', lambda e, idx=class_index: self.update_from_entry(idx))
        
        self.entry_widgets[class_index] = entry
        
        count_label = ttk.Label(control_frame, textvariable=count_var, width=4,
                              anchor="center", background="white", relief="solid")
        count_label.pack(side="left", padx=2)
        self.count_labels[class_index] = count_label
        
        ttk.Button(control_frame, text="+", width=2,
                  command=lambda idx=class_index: self.increment(idx)).pack(side="left")
        
        # Đồng bộ giá trị
        self.sync_entry_value(class_index)
        
        return {
            'class_name': class_name,
            'class_index': class_index,
            'is_others_section': is_others_section
        }

    def _rebuild_ui_from_cache(self, cached_data):
        """Xây dựng lại UI từ cache - rất nhanh"""
        # Xóa widget cũ
        for widget in self.class_frame.winfo_children():
            widget.destroy()

        self.count_labels.clear()
        self.entry_widgets.clear()

        # Tạo layout 3 cột
        columns_frame = ttk.Frame(self.class_frame)
        columns_frame.pack(fill="both", expand=True)
        
        left_column = ttk.Frame(columns_frame)
        left_column.pack(side="left", fill="both", expand=True, padx=5)
        
        center_column = ttk.Frame(columns_frame)
        center_column.pack(side="left", fill="both", expand=True, padx=5)
        
        right_column = ttk.Frame(columns_frame)
        right_column.pack(side="left", fill="both", expand=True, padx=5)

        # Xây dựng từ cache
        for section_data in cached_data['left']:
            self._build_section_from_cache(left_column, section_data)
        for section_data in cached_data['center']:
            self._build_section_from_cache(center_column, section_data)
        for section_data in cached_data['right']:
            self._build_section_from_cache(right_column, section_data)

    def _build_section_from_cache(self, column, section_data):
        """Xây dựng section từ cache"""
        section_frame = ttk.LabelFrame(column, text=section_data['parent_name'], padding=5)
        section_frame.pack(fill="both", expand=True, pady=5)
        
        for child_data in section_data['children']:
            self._create_class_row(
                section_frame, 
                child_data['class_name'], 
                child_data['class_index'], 
                child_data['is_others_section']
            )

    # ----------------------------
    # CÁC HÀM CÒN LẠI - ĐÃ SỬA LỖI
    # ----------------------------
    def sync_entry_value(self, class_index):
        if class_index in self.entry_widgets:
            self.entry_widgets[class_index].delete(0, tk.END)
            self.entry_widgets[class_index].insert(0, str(self.counts_all[class_index].get()))

    def update_from_entry(self, class_index):
        if class_index in self.entry_widgets:
            entry_value = self.entry_widgets[class_index].get()
            if entry_value.strip() == "":
                new_value = 0
            else:
                try:
                    new_value = int(entry_value)
                    if new_value < 0:
                        new_value = 0
                        self.entry_widgets[class_index].delete(0, tk.END)
                        self.entry_widgets[class_index].insert(0, "0")
                except ValueError:
                    new_value = self.counts_all[class_index].get()
                    self.entry_widgets[class_index].delete(0, tk.END)
                    self.entry_widgets[class_index].insert(0, str(new_value))
            
            self.counts_all[class_index].set(new_value)

    def increment(self, class_index):
        self.counts_all[class_index].set(self.counts_all[class_index].get() + 1)
        self.sync_entry_value(class_index)

    def decrement(self, class_index):
        val = self.counts_all[class_index].get()
        if val > 0:
            self.counts_all[class_index].set(val - 1)
            self.sync_entry_value(class_index)

    def start_timer(self):
        if self.is_paused:
            self.session_start = time.time()
            self.is_paused = False

    def pause_timer(self):
        if not self.is_paused and self.session_start:
            elapsed = time.time() - self.session_start
            self.total_work_time += elapsed
            self.is_paused = True
            self.session_start = None

    def get_total_elapsed_seconds(self):
        if self.is_paused or not self.session_start:
            return int(self.total_work_time)
        else:
            current_elapsed = time.time() - self.session_start
            return int(self.total_work_time + current_elapsed)

    def format_seconds_hms(self, seconds):
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    def update_timer_display(self):
        total_seconds = self.get_total_elapsed_seconds()
        self.timer_label.config(text=f"Working Time: {self.format_seconds_hms(total_seconds)}")
        self.root.after(1000, self.update_timer_display)

    def save_to_excel(self):
        dataset_name = self.dataset_name_entry.get().strip()
        if not dataset_name:
            messagebox.showerror("Error", "Please enter dataset name")
            return

        save_lang = tk.StringVar(value=self.current_language.get())
        save_window = tk.Toplevel(self.root)
        save_window.title("Select Save Language")

        ttk.Label(save_window, text="Select language to save:").pack(pady=10)
        lang_menu = ttk.OptionMenu(save_window, save_lang, self.current_language.get(), *self.languages)
        lang_menu.pack(pady=5)

        def confirm_save():
            self._save_to_excel_internal(dataset_name, save_lang.get())
            save_window.destroy()

        ttk.Button(save_window, text="Save", command=confirm_save).pack(pady=10)

    def _save_to_excel_internal(self, dataset_name, save_lang):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        elapsed_seconds = self.get_total_elapsed_seconds()
        elapsed_str = self.format_seconds_hms(elapsed_seconds)

        # CHỈ LẤY ĐÚNG 67 CLASS CON
        save_classes = []
        save_counts = []
        
        # Tạo danh sách tất cả class con theo thứ tự
        all_child_indices = []
        for parent_index in sorted(self.parent_classes.keys()):
            child_indices = self._find_child_indices(parent_index)
            all_child_indices.extend(child_indices)
        
        # Sắp xếp và chỉ lấy 67 class con
        all_child_indices = sorted(all_child_indices)
        
        # In debug info
        print(f"DEBUG: Total child indices found: {len(all_child_indices)}")
        print(f"DEBUG: Child indices: {all_child_indices}")
        
        for index in all_child_indices:
            if index in self.class_mapping:
                class_name = self.class_mapping[index].get(save_lang, f"Class_{index}")
                save_classes.append(class_name)
                save_counts.append(self.counts_all[index].get())

        print(f"DEBUG: Saving {len(save_classes)} classes to Excel")

        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            sheet_name = f"Counts_{save_lang}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Counts_{save_lang}"

        # Tạo header nếu chưa có hoặc không khớp
        if ws.max_row == 0:
            ws.append(["Dataset Name", "Working Time"] + save_classes)
        else:
            # Kiểm tra header hiện tại
            current_header = [cell.value for cell in ws[1]]
            expected_header = ["Dataset Name", "Working Time"] + save_classes
            
            if len(current_header) != len(expected_header) or current_header != expected_header:
                # Header không khớp, xóa toàn bộ sheet và tạo lại
                ws.delete_rows(1, ws.max_row)
                ws.append(["Dataset Name", "Working Time"] + save_classes)

        # Tìm và cập nhật dòng hiện có, hoặc thêm mới
        found = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row[0].value == dataset_name:
                # Cập nhật dòng hiện có
                row[0].value = dataset_name
                row[1].value = elapsed_str
                for i, cell in enumerate(row[2:2+len(save_counts)]):
                    if i < len(save_counts):
                        cell.value = save_counts[i]
                # Xóa dữ liệu thừa nếu có
                for i in range(2+len(save_counts), len(row)):
                    if i < len(row):
                        row[i].value = None
                found = True
                break

        if not found:
            # Thêm dòng mới
            ws.append([dataset_name, elapsed_str] + save_counts)

        wb.save(filename)
        messagebox.showinfo("Success", f"Saved to {filename}\nTotal classes: {len(save_classes)}")

    def load_from_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        dataset_name = self.dataset_name_entry.get().strip()
        if not dataset_name:
            messagebox.showerror("Error", "Please enter dataset name before loading")
            return

        try:
            wb = openpyxl.load_workbook(filename)
            
            ws = None
            loaded_lang = None
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                header = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
                if len(header) >= 3:
                    excel_classes = header[2:]
                    for lang in self.languages:
                        # Tạo danh sách class con cho ngôn ngữ này
                        lang_child_classes = []
                        all_child_indices = []
                        for parent_index in sorted(self.parent_classes.keys()):
                            child_indices = self._find_child_indices(parent_index)
                            all_child_indices.extend(child_indices)
                        
                        all_child_indices = sorted(all_child_indices)
                        
                        for index in all_child_indices:
                            if index in self.class_mapping:
                                class_name = self.class_mapping[index].get(lang, f"Class_{index}")
                                lang_child_classes.append(class_name)
                        
                        if excel_classes == lang_child_classes:
                            ws = sheet
                            loaded_lang = lang
                            break
                if ws:
                    break
            
            if not ws:
                ws = wb.active

            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] == dataset_name:
                    time_str = str(row[1])
                    if ':' in time_str:
                        h, m, s = map(int, time_str.split(":"))
                        self.total_work_time = h*3600 + m*60 + s
                    self.session_start = None
                    self.is_paused = True

                    header = [cell.value for cell in ws[1]]
                    excel_classes = header[2:]
                    
                    for index in self.class_indexes:
                        self.counts_all[index].set(0)
                    
                    # Tạo danh sách class con để load
                    all_child_indices = []
                    for parent_index in sorted(self.parent_classes.keys()):
                        child_indices = self._find_child_indices(parent_index)
                        all_child_indices.extend(child_indices)
                    
                    all_child_indices = sorted(all_child_indices)
                    
                    for i, excel_class in enumerate(excel_classes):
                        if i + 2 < len(row) and row[i + 2] is not None:
                            # Tìm class index tương ứng
                            for index in all_child_indices:
                                if index in self.class_mapping:
                                    lang_dict = self.class_mapping[index]
                                    if excel_class in lang_dict.values():
                                        self.counts_all[index].set(row[i + 2])
                                        if index in self.entry_widgets:
                                            self.sync_entry_value(index)
                                        break

                    found = True
                    if loaded_lang:
                        self.current_language.set(loaded_lang)
                        self.update_language(loaded_lang)
                    
                    messagebox.showinfo("Success", f"Loaded dataset '{dataset_name}' from {filename}")
                    break

            if not found:
                messagebox.showinfo("Info", f"Dataset '{dataset_name}' not found in {filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {e}")

# ==========================
# CHẠY CHƯƠNG TRÌNH
# ==========================
if __name__ == "__main__":
    root = tk.Tk()
    app = CounterApp(root)
    root.mainloop()